# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


#print('savan')

#price = 10
#price = 20
#print(price)

#name = input('what is your name? ')
#print(' Hi ' + name)

#course = 'python for biginners'
#print(course[0])
#print(course[-1])
#print(course[0:3])

#print(course.upper())

#x = 10 ** 3
#print(x)

#x = 10 + 2 * 5
#print(x)
"""
is_hot = False
is_cold = True

if is_hot:
   print("It's a hot day")
elif is_cold:
   print("It's cold day")
else:
    print("It's lovely day")
"""
"""
i = 6

while i>0:
    print('*' * i)
    i -= 1
print("Done")
"""
"""
command = ""
started = False
while True:
    command = input('> ').lower()
    if command == "start":
        if started:
            print("car is aiready started")
        else:
            started = True
            print('car is start......')
    elif command == "stop":
        if not started:
            print('car is already stopped')
        else:
            started = False
            print("car stopped")
    elif command == "help":
        print( """
#start - to start car
#stop - to stop car
#quit - to quit
 #       """ )
  #  elif command == "quit":
   #     break
    #else:
     #   print("Sorry, I don't understand ")
"""
"""
#for x in range(0,10):
 #   print(x)
"""
for x in range (4):
    for y in range (3):
        print(f"({x},{y})")


"""

#number = [5,2,5,2,5]
#for x in number:
 #   print('x' * x)
"""
list = [45,65,78,2,46,99,34,54,23,1000]
max = list[0]
for x in list:
    if x > max:
        max = x
print(max)
"""
"""
number = [3,6,9,43,5]
number.append(23)
print(number)
number.insert(3,56)
print(number)
number.remove(6)
print(number)
number.pop()
print(number)
"""
"""
#list
numbers = [4,5,6,2,3,2,5,6,7,9,8,7,9,1]
unique = []
for number in numbers:
    if number not in unique:
        unique.append(number)
print(unique)

"""
"""
#tuples
numbers = (1,2,3)
numbers[0] = 10
print(numbers[0])
"""
"""
#unpacking
coordinates = (1,2,3)
x,y,z = coordinates
print(x)
print(y)
print(z) 
"""

"""
#dictionary
customer = {
    "name" : "ABC XYZ",
    "age" : 39,
    "mno" : "1234567890"
}
print(customer["name"])



phone = input("Phone: ")
digit_mapping = {
    "1" : "One",
    "2" : "Two",
    "3" : "Three",
    "4" : "Four"
}
output = ""
for x in phone:
    output += digit_mapping.get( x , "!" ) + ""
print(output)
"""
"""
#emoji converter using dictionary

message = input(">")
words = message.split(' ')
emojis = {
    ":)": "🙂",
    "(:": "😢"
}
output = ""
for word in words:
    output += emojis.get(word, word) + " "
print(output)
"""

"""
#function

def abc():
    print("Hiiii")
    print("How are you")


print("Start")
abc()
print("Stop")
"""
"""
#parameter

def abc(first_name,last_name):
    print(f"Hiiii {first_name} {last_name}")
    print("How are you")


print("Start")
abc("savan",last_name="patel")
print("Stop")
"""
"""
#retunr statemant
def square(num):
    return num * num


print(square(4))
"""
"""
#create a reuseable function
def emojis_converter(message):
    words = message.split(' ')
    emojis = {
        ":)": "🙂",
        "(:": "😢"
    }
    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
        return output
message = input(">")
result = emojis_converter(message)
print(result)
"""
"""
#exception
try:
    age = int(input("Age: "))
    income = 10000
    risk = income/age
    print(risk)
except ZeroDivisionError:
    print("Age cannot be 0")
except ValueError:
    print("Invalid value")
"""
"""
#class
class Point:
    def move(self):
        print("move")

    def draw(self):
        print("draw")

p1 = Point()
p1.draw()
p1.x=10
p1.y=20
print(p1.x,p1.y)

p2 = Point()
p2.move()
"""
"""
#constructor

class Person:
    def __init__(self, name):
        self.name = name

    def talk(self):
        print(f"I'm {self.name}")


p1 = Person("abc xyz")
print(p1.name,p1.talk())
"""
"""
#inheritance

class Top:
    def walk(self):
        print("walk")


class Dog(Top):
    def bark(self):
        print("Bark")


class Cat(Top):
    pass

d1 = Dog()
d1.bark()

c1 = Cat()
c1.walk()
"""
"""
#modules
import converter

print(converter.kg_to_lbs(20))
"""
"""
import utils

numbers = [53,45,23,76]

max = utils.find_max(numbers)
print(max)
"""

#packages


"""
#module
import random

name = ['abc','xyz','qwe','tfhv']
leader = random.choice(name)
print(leader)

"""
"""
#module
import random
class Dice:
    def roll(self):
        first = random.randint(1,6)
        second = random.randint(1,6)
        return first,second


dice = Dice()
print(dice.roll())
"""

"""
#files and directory

from pathlib import Path

path = Path()
for file in path.glob('*.py'):
    print(file)
"""
#pypi and pip

"""
#open terminal and input to command like pip install openpyxl

import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb = xl.load_workbook('transaction.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1,1)

for row in range(2,sheet.max_row + 1):
    cell = sheet.cell(row,3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=4,
                   max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'f2')
wb.save('transaction1.xlsx ')
"""

